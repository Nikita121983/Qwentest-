"""
Быстрый поиск и замена формул
Читает строками через iter_rows, не по ячейке
"""
import openpyxl
import re
import os
import shutil
import time
from collections import defaultdict

PROCESS_FILE = r'c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Обработка.xlsx'

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")

def unwrap_formula(formula, indent=0):
    """Разворачивает вложенные IF в читаемый вид"""
    lines = []
    depth = 0
    current = []
    i = 0
    
    while i < len(formula):
        chunk = ''
        # Проверяем функции
        func_match = re.match(r'([A-Z_.]+)\(', formula[i:])
        if func_match:
            depth += 1
            current.append(formula[i:i+len(func_match.group(0))])
            current.append('\n' + '  ' * depth)
            i += len(func_match.group(0))
        elif formula[i] == ',':
            if depth > 1:
                current.append(',')
                current.append('\n' + '  ' * depth)
            else:
                current.append(', ')
            i += 1
        elif formula[i] == ')':
            depth -= 1
            current.append(')')
            i += 1
        else:
            # Читаем до следующей скобки/запятой/функции
            j = i
            while j < len(formula) and formula[j] not in '(),' and not re.match(r'[A-Z_.]+\(', formula[j:]):
                j += 1
            current.append(formula[i:j])
            i = j
    
    return ''.join(current)

def get_all_formulas():
    """Быстро извлекает ВСЕ формулы из файла"""
    log(f"Открываю {os.path.basename(PROCESS_FILE)}...")
    wb = openpyxl.load_workbook(PROCESS_FILE, data_only=False)
    
    all_formulas = {}
    all_flat = []
    
    for ws in wb.worksheets:
        ws_name = ws.title
        ws_formulas = {}
        max_col = min(ws.max_column, 225)
        
        # Читаем весь диапазон одной итерацией
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 502), max_col=max_col, values_only=False):
            for c in row:
                if c.value and isinstance(c.value, str) and c.value.startswith('='):
                    col_letter = openpyxl.utils.get_column_letter(c.column)
                    tpl = re.sub(r'\$?[A-Z]+\$?\d+', '$REF', c.value)
                    
                    if col_letter not in ws_formulas:
                        ws_formulas[col_letter] = {}
                    if tpl not in ws_formulas[col_letter]:
                        ws_formulas[col_letter][tpl] = {
                            'count': 0, 'formula': c.value, 'rows': [], 'first': c.value
                        }
                    ws_formulas[col_letter][tpl]['count'] += 1
                    if len(ws_formulas[col_letter][tpl]['rows']) < 5:
                        ws_formulas[col_letter][tpl]['rows'].append(c.row)
                    all_flat.append((ws_name, c.coordinate, c.value))
        
        all_formulas[ws_name] = ws_formulas
    
    wb.close()
    log(f"Извлечено {len(all_flat):,} формул из {len(all_formulas)} листов")
    return all_formulas, all_flat

def show_sheet_summary(sheet_name, all_formulas):
    """Сводка по листу"""
    if sheet_name not in all_formulas:
        print(f"Лист '{sheet_name}' не найден")
        return
    
    cols = all_formulas[sheet_name]
    print(f"\n{'='*80}")
    print(f"{sheet_name}: {len(cols)} столбцов с формулами")
    print(f"{'='*80}")
    
    for col_letter, variants in sorted(cols.items()):
        total = sum(v['count'] for v in variants.values())
        n_variants = len(variants)
        marker = " 🔀" if n_variants > 1 else ""
        first_formula = list(variants.values())[0]['formula'][:80]
        
        print(f"  {col_letter:4s} {total:>5} ф-л, {n_variants} шабл(ов){marker}  {first_formula}")

def find_formulas(keyword, all_flat, max_results=30):
    """Быстрый поиск по всем формулам"""
    matches = []
    for sheet, cell, formula in all_flat:
        if keyword.lower() in formula.lower():
            matches.append((sheet, cell, formula))
    
    print(f"\n{'='*80}")
    print(f"Найдено {len(matches)} совпадений для '{keyword}'")
    print(f"{'='*80}")
    
    # Группируем по листу и столбцу
    by_sheet_col = defaultdict(list)
    for sheet, cell, formula in matches:
        col = re.match(r'([A-Z]+)', cell).group(1)
        by_sheet_col[(sheet, col)].append((cell, formula))
    
    shown = 0
    for (sheet, col), cells in sorted(by_sheet_col.items()):
        n_variants = len(set(re.sub(r'\$?[A-Z]+\$?\d+', '$REF', f) for _, f in cells))
        print(f"\n  {sheet}!{col} ({len(cells)} ячеек, {n_variants} вариант(ов))")
        
        if shown < max_results:
            first_cell, first_formula = cells[0]
            print(f"    {first_cell}: {first_formula[:120]}")
            if len(first_formula) > 80:
                print(f"    Развернутая:")
                for line in unwrap_formula(first_formula).split('\n'):
                    if line.strip():
                        print(f"      {line}")
            shown += 1
    
    if len(matches) > max_results:
        print(f"\n  ... и ещё {len(matches) - max_results} совпадений")
    
    return matches

def show_column_variants(sheet_name, col_letter, all_formulas):
    """Все варианты формул в столбце"""
    if sheet_name not in all_formulas:
        print(f"Лист '{sheet_name}' не найден")
        return
    if col_letter not in all_formulas[sheet_name]:
        print(f"Столбец '{col_letter}' не найден в '{sheet_name}'")
        return
    
    variants = all_formulas[sheet_name][col_letter]
    total = sum(v['count'] for v in variants.values())
    
    print(f"\n{'='*80}")
    print(f"{sheet_name}!{col_letter} — {len(variants)} варианта(ов), всего {total} ячеек")
    print(f"{'='*80}")
    
    for i, (tpl, info) in enumerate(variants.items()):
        print(f"\n--- Вариант {i+1}: {info['count']} ячеек (строки {info['rows']}) ---")
        print(f"Формула: {info['formula'][:150]}")
        
        if len(info['formula']) > 80:
            print(f"\nРазвернутая:")
            for line in unwrap_formula(info['formula']).split('\n'):
                if line.strip():
                    print(f"  {line}")

def replace_formula(sheet_name, col_letter, keyword, new_formula, dry_run=True):
    """Заменяет формулы"""
    if dry_run:
        log(f"[DRY RUN] Замена в {sheet_name}!{col_letter} '{keyword}' → ...")
    else:
        log(f"Замена в {sheet_name}!{col_letter} '{keyword}' → '{new_formula[:60]}'")
    
    wb = openpyxl.load_workbook(PROCESS_FILE)
    ws = wb[sheet_name]
    col_num = openpyxl.utils.column_index_from_string(col_letter)
    
    count = 0
    for row in range(1, min(ws.max_row + 1, 502)):
        cell = ws.cell(row=row, column=col_num)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            if keyword.lower() in cell.value.lower():
                # Адаптируем формулу для строки
                new = re.sub(
                    r'(\$?)([A-Z]+)(\$?)(\d+)',
                    lambda m: m.group(1) + m.group(2) + m.group(3) + str(cell.row),
                    new_formula
                )
                
                if dry_run:
                    print(f"  {cell.coordinate}: {cell.value[:80]}")
                    print(f"    → {new[:80]}")
                else:
                    cell.value = new
                
                count += 1
    
    if not dry_run:
        backup = PROCESS_FILE + '.bak.' + time.strftime('%Y%m%d_%H%M%S')
        shutil.copy2(PROCESS_FILE, backup)
        print(f"\n  Бэкап: {backup}")
        wb.save(PROCESS_FILE)
        print(f"  ✅ Обновлено {count} ячеек")
    
    wb.close()
    return count

# ============================================================================
# MAIN — быстрый запуск
# ============================================================================
if __name__ == '__main__':
    all_formulas, all_flat = get_all_formulas()
    
    print(f"\n{'='*80}")
    print("ИНСТРУМЕНТ ГОТОВ. Доступные команды:")
    print("="*80)
    print("""
# Сводка по листу:
show_sheet_summary('Расчет', all_formulas)
show_sheet_summary('Замер', all_formulas)

# Найти формулу:
find_formulas('Словарь условий', all_flat)
find_formulas('INDEX', all_flat)
find_formulas('IFERROR', all_flat)

# Варианты в столбце:
show_column_variants('Расчет', 'AV', all_formulas)
show_column_variants('Замер', 'D', all_formulas)

# Заменить (dry run — покажет что будет):
replace_formula('Расчет', 'AV', 'старый_текст', 'новая_формула', dry_run=True)

# Заменить (реально — с бэкапом):
replace_formula('Расчет', 'AV', 'старый_текст', 'новая_формула', dry_run=False)
""")
    
    # Авто-показ: сводка по ключевым листам
    for sheet in ['Расчет', 'Замер', 'Список фильтр', 'Маркировка']:
        show_sheet_summary(sheet, all_formulas)
