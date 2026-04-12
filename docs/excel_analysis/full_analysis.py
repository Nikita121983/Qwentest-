"""
Полный анализ Excel-конвейера обработки замеров дверей.
Извлекает ВСЕ формулы, строит карту зависимостей, находит варианты и артефакты.
"""
import openpyxl
import re
import json
from collections import defaultdict, Counter
import os
import time

PROCESS_FILE = r'c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Обработка.xlsx'
SOURCE_FILE = r'c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Юнтолово 8 корпус 1-22 этажи Квартиры Список измерений edit.xlsx'
OUTPUT_DIR = r'C:\Users\Nik\SyncthingServiceAcct\Qwentest\analysis'

os.makedirs(OUTPUT_DIR, exist_ok=True)

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")

# ============================================================================
# 1. ИЗВЛЕЧЕНИЕ ВСЕХ ФОРМУЛ
# ============================================================================
def extract_all_formulas(filepath):
    """Извлекает ВСЕ формулы из всех листов файла"""
    wb = openpyxl.load_workbook(filepath, data_only=False, read_only=False)
    result = {}
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        sheet_data = {
            'formulas': {},      # {coord: formula}
            'values': {},        # {coord: value}
            'formula_cols': defaultdict(list),  # {col_letter: [(row, formula)]}
            'unique_formulas': {},  # {formula_template: [coords]}
            'total_cells': 0,
            'formula_count': 0,
        }
        
        log(f"  Читаю лист '{ws_name}' ({ws.max_row}x{ws.max_column})...")
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            for cell in row:
                sheet_data['total_cells'] += 1
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    col_letter = re.match(r'([A-Z]+)', cell.coordinate).group(1)
                    sheet_data['formulas'][cell.coordinate] = formula
                    sheet_data['formula_cols'][col_letter].append((cell.row, formula))
                    sheet_data['formula_count'] += 1
                    
                    # Нормализуем формулу для группировки (заменяем ссылки на ячейки)
                    template = re.sub(r'\$?[A-Z]+\$?\d+', '$CELL', formula)
                    if template not in sheet_data['unique_formulas']:
                        sheet_data['unique_formulas'][template] = []
                    sheet_data['unique_formulas'][template].append(cell.coordinate)
                    
                elif cell.value is not None:
                    sheet_data['values'][cell.coordinate] = cell.value
        
        result[ws_name] = sheet_data
    
    wb.close()
    return result

# ============================================================================
# 2. КАРТА ЗАВИСИМОСТЕЙ МЕЖДУ ЛИСТАМИ
# ============================================================================
def build_dependency_map(formulas_data):
    """Строит карту: какой лист от какого зависит"""
    deps = defaultdict(lambda: defaultdict(int))  # {from_sheet: {to_sheet: count}}
    ext_file_refs = defaultdict(lambda: defaultdict(int))  # {sheet: {ext_ref: count}}
    
    for ws_name, data in formulas_data.items():
        for coord, formula in data['formulas'].items():
            # Ссылки на другие листы: 'SheetName'!A1 или SheetName!A1
            sheet_refs = re.findall(r"'?([A-Za-zА-Яа-я0-9_ ()]+)'?!", formula)
            for ref in sheet_refs:
                if ref.strip() and ref.strip() != ws_name:
                    deps[ws_name][ref.strip()] += 1
            
            # Внешние ссылки на файлы: [1] или [filename.xlsx]
            ext_refs = re.findall(r'\[([^\]]+)\]', formula)
            for ref in ext_refs:
                ext_file_refs[ws_name][ref] += 1
    
    return dict(deps), dict(ext_file_refs)

# ============================================================================
# 3. АНАЛИЗ ВАРИАНТОВ ФОРМУЛ В СТОЛБЦАХ
# ============================================================================
def analyze_column_variants(formulas_data):
    """Находит столбцы с несколькими вариантами формул"""
    result = {}
    
    for ws_name, data in formulas_data.items():
        col_variants = {}
        for col_letter, cell_formulas in data['formula_cols'].items():
            if len(cell_formulas) < 2:
                continue
            
            # Группируем по шаблону формулы
            templates = defaultdict(list)
            for row, formula in cell_formulas:
                template = re.sub(r'\$?[A-Z]+\$?\d+', '$CELL', formula)
                templates[template].append((row, formula))
            
            if len(templates) > 1:
                col_variants[col_letter] = {
                    'total_cells': len(cell_formulas),
                    'variants': {
                        t: {
                            'count': len(coords),
                            'example': coords[0][1][:150],
                            'rows': [r for r, _ in coords[:5]]
                        }
                        for t, coords in templates.items()
                    }
                }
        
        if col_variants:
            result[ws_name] = col_variants
    
    return result

# ============================================================================
# 4. ПОИСК МЁРТВЫХ ЗОН (формулы ссылаются на пустые данные, неиспользуемые столбцы)
# ============================================================================
def find_dead_zones(formulas_data):
    """Находит столбцы с формулами, которые ссылаются на пустые ячейки или неиспользуемы"""
    result = {}
    
    for ws_name, data in formulas_data.items():
        dead_cols = []
        
        for col_letter, cell_formulas in data['formula_cols'].items():
            if not cell_formulas:
                continue
            
            # Проверяем, есть ли результаты (values) в этом столбце
            value_count = sum(1 for coord in data['values'] 
                            if re.match(rf'{col_letter}\d+', coord))
            
            # Проверяем, ссылается ли кто-то на этот столбец
            referenced = False
            for other_ws, other_data in formulas_data.items():
                if other_ws == ws_name:
                    continue
                for coord, formula in other_data['formulas'].items():
                    if col_letter in formula:
                        referenced = True
                        break
                if referenced:
                    break
            
            # Формулы есть, но значений нет и никто не ссылается
            if data['formula_count'] > 0 and value_count == 0 and not referenced:
                dead_cols.append({
                    'column': col_letter,
                    'formulas': len(cell_formulas),
                    'values': value_count,
                    'referenced': referenced
                })
        
        if dead_cols:
            result[ws_name] = dead_cols
    
    return result

# ============================================================================
# 5. СТАТИСТИКА ПО УНИКАЛЬНЫМ ФОРМУЛАМ
# ============================================================================
def analyze_formula_patterns(formulas_data):
    """Анализирует паттерны формул"""
    all_templates = defaultdict(list)  # {template: [(sheet, coord)]}
    
    for ws_name, data in formulas_data.items():
        for coord, formula in data['formulas'].items():
            template = re.sub(r'\$?[A-Z]+\$?\d+', '$CELL', formula)
            all_templates[template].append((ws_name, coord))
    
    # Сортируем по количеству повторений
    sorted_templates = sorted(all_templates.items(), key=lambda x: len(x[1]), reverse=True)
    
    return sorted_templates[:100]  # Топ-100 самых повторяющихся

# ============================================================================
# ОСНОВНОЙ ПРОЦЕСС
# ============================================================================
log("=" * 80)
log("ШАГ 1: Извлечение всех формул из Обработка.xlsx")
log("=" * 80)
formulas_data = extract_all_formulas(PROCESS_FILE)

log("\n" + "=" * 80)
log("ШАГ 2: Карта зависимостей")
log("=" * 80)
dep_map, ext_refs = build_dependency_map(formulas_data)

log("\n" + "=" * 80)
log("ШАГ 3: Анализ вариантов формул в столбцах")
log("=" * 80)
col_variants = analyze_column_variants(formulas_data)

log("\n" + "=" * 80)
log("ШАГ 4: Поиск мёртвых зон")
log("=" * 80)
dead_zones = find_dead_zones(formulas_data)

log("\n" + "=" * 80)
log("ШАГ 5: Паттерны формул")
log("=" * 80)
top_patterns = analyze_formula_patterns(formulas_data)

# ============================================================================
# ИТОГОВЫЙ ОТЧЁТ
# ============================================================================
log("\n" + "=" * 80)
log("ИТОГОВЫЙ ОТЧЁТ")
log("=" * 80)

# Общая статистика
print("\n--- ОБЩАЯ СТАТИСТИКА ---")
total_formulas = sum(d['formula_count'] for d in formulas_data.values())
total_cells = sum(d['total_cells'] for d in formulas_data.values())
total_unique_templates = sum(len(d['unique_formulas']) for d in formulas_data.values())

print(f"Всего ячеек: {total_cells:,}")
print(f"Всего формул: {total_formulas:,}")
print(f"Всего уникальных шаблонов формул: {total_unique_templates:,}")

for ws_name, data in sorted(formulas_data.items(), key=lambda x: x[1]['formula_count'], reverse=True):
    print(f"  {ws_name:45s} формул: {data['formula_count']:>7,}  уникальных шаблонов: {len(data['unique_formulas']):>4}  ячеек: {data['total_cells']:>7,}")

# Карта зависимостей
print(f"\n--- КАРТА ЗАВИСИМОСТЕЙ ({len(dep_map)} листов с зависимостями) ---")
for from_sheet, targets in sorted(dep_map.items()):
    for to_sheet, count in sorted(targets.items(), key=lambda x: x[1], reverse=True)[:10]:
        print(f"  {from_sheet:45s} -> {to_sheet:45s}  ({count:,} ссылок)")

# Внешние ссылки
print(f"\n--- ВНЕШНИЕ ССЫЛКИ НА ФАЙЛЫ ---")
for sheet, refs in ext_refs.items():
    for ref, count in refs.items():
        print(f"  {sheet:45s} -> [{ref}] ({count:,} формул)")

# Варианты формул
print(f"\n--- СТОЛБЦЫ С НЕСКОЛЬКИМИ ВАРИАНТАМИ ФОРМУЛ ---")
for ws_name, cols in col_variants.items():
    for col, info in cols.items():
        print(f"\n  {ws_name}!{col} ({info['total_cells']} ячеек, {len(info['variants'])} варианта):")
        for i, (template, data) in enumerate(info['variants'].items()):
            print(f"    Вариант {i+1}: {data['count']} ячеек, строки {data['rows']}")
            print(f"      {data['example'][:100]}")

# Мёртвые зоны
print(f"\n--- МЁРТВЫЕ ЗОНЫ (формулы без значений и без ссылок) ---")
for ws_name, cols in dead_zones.items():
    for col in cols:
        print(f"  {ws_name}!{col['column']}: {col['formulas']} формул, {col['values']} значений, referenced={col['referenced']}")

# Топ паттернов
print(f"\n--- ТОП-30 САМЫХ ПОВТОРЯЮЩИХСЯ ФОРМУЛ ---")
for i, (template, locations) in enumerate(top_patterns[:30]):
    sheets = Counter(ws for ws, _ in locations)
    print(f"  #{i+1}: {len(locations):>6,} повторений на листах: {dict(sheets)}")
    print(f"      {template[:120]}")

# ============================================================================
# СОХРАНЕНИЕ В ФАЙЛЫ
# ============================================================================
log("\n--- Сохранение результатов ---")

# 1. Полная статистика
stats = {
    'summary': {
        'total_cells': total_cells,
        'total_formulas': total_formulas,
        'total_unique_templates': total_unique_templates,
        'sheets': {ws: {
            'formulas': d['formula_count'],
            'unique_templates': len(d['unique_formulas']),
            'total_cells': d['total_cells']
        } for ws, d in formulas_data.items()}
    },
    'dependencies': {ws: dict(targets) for ws, targets in dep_map.items()},
    'external_refs': {ws: dict(refs) for ws, refs in ext_refs.items()},
    'column_variants': {
        ws: {col: {
            'total': info['total_cells'],
            'variant_count': len(info['variants']),
            'variants': {t: v['count'] for t, v in info['variants'].items()}
        } for col, info in cols.items()}
        for ws, cols in col_variants.items()
    },
    'dead_zones': dead_zones,
    'top_formula_patterns': [
        {'template': t[:200], 'count': len(locs), 'sheets': dict(Counter(ws for ws, _ in locs))}
        for t, locs in top_patterns[:50]
    ]
}

with open(os.path.join(OUTPUT_DIR, 'full_analysis.json'), 'w', encoding='utf-8') as f:
    json.dump(stats, f, ensure_ascii=False, indent=2)

# 2. Список всех уникальных формул
with open(os.path.join(OUTPUT_DIR, 'all_formulas.txt'), 'w', encoding='utf-8') as f:
    for template, locations in top_patterns:
        f.write(f"\n{'='*80}\n")
        f.write(f"Шаблон ({len(locations)} повторений): {template[:200]}\n")
        f.write(f"Листа: {dict(Counter(ws for ws, _ in locations))}\n")
        for ws, coord in locations[:3]:
            f.write(f"  {ws}!{coord}\n")

# 3. Карта зависимостей (DOT формат для Graphviz)
with open(os.path.join(OUTPUT_DIR, 'dependencies.dot'), 'w', encoding='utf-8') as f:
    f.write("digraph dependencies {\n")
    f.write("  rankdir=LR;\n")
    f.write("  node [shape=box, style=rounded];\n\n")
    
    for from_sheet, targets in dep_map.items():
        for to_sheet, count in targets.items():
            if count > 10:  # Только значимые связи
                width = max(1, min(10, count / 1000))
                f.write(f'  "{from_sheet}" -> "{to_sheet}" [label="{count:,}", penwidth={width:.1f}];\n')
    
    f.write("}\n")

log(f"\nРезультаты сохранены в {OUTPUT_DIR}/")
log("ГОТОВО!")
