"""
Супер-быстрый анализ: читаем строками через iter_rows()
"""
import openpyxl
import json
import time
import os

PROCESS_FILE = r'c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Обработка.xlsx'
OUTPUT_DIR = r'C:\Users\Nik\SyncthingServiceAcct\Qwentest\analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")

def analyze():
    log("Открываю оба файла одновременно...")
    t0 = time.time()
    wb_form = openpyxl.load_workbook(PROCESS_FILE, data_only=False)
    wb_val = openpyxl.load_workbook(PROCESS_FILE, data_only=True)
    log(f"Файлы открыты за {time.time()-t0:.1f}с")
    
    results = {}
    
    for ws_name in wb_form.sheetnames:
        ws_form = wb_form[ws_name]
        ws_val = wb_val[ws_name]
        
        log(f"Анализирую '{ws_name}' ({ws_form.max_row}x{ws_form.max_column})...")
        
        sheet_info = {
            'formula_cells': 0,
            'formula_with_value': 0,
            'formula_empty': 0,
            'formula_error': 0,
            'sample_results': [],
            'dead_columns': [],
            'live_columns': [],
            'mixed_columns': [],
            'column_details': {}
        }
        
        col_stats = {}
        max_rows = min(ws_form.max_row, 502)
        max_cols = min(ws_form.max_column, 225)
        
        # Читаем СТРОКАМИ — быстро
        for row_num in range(1, max_rows + 1):
            form_row = list(ws_form.iter_rows(min_row=row_num, max_row=row_num, 
                                               max_col=max_cols, values_only=False))[0]
            val_row = list(ws_val.iter_rows(min_row=row_num, max_row=row_num, 
                                             max_col=max_cols, values_only=False))[0]
            
            for col_num, (fc, vc) in enumerate(zip(form_row, val_row), 1):
                if fc.value and isinstance(fc.value, str) and fc.value.startswith('='):
                    col_letter = openpyxl.utils.get_column_letter(col_num)
                    result = vc.value
                    
                    sheet_info['formula_cells'] += 1
                    
                    if col_letter not in col_stats:
                        col_stats[col_letter] = {
                            'total': 0, 'with_value': 0, 'empty': 0,
                            'sample_formula': fc.value[:100],
                            'sample_result': None
                        }
                    
                    cs = col_stats[col_letter]
                    cs['total'] += 1
                    
                    if result is None or result == "":
                        sheet_info['formula_empty'] += 1
                        cs['empty'] += 1
                    elif isinstance(result, str) and (result.startswith('#') or result.startswith('!')):
                        sheet_info['formula_error'] += 1
                        cs['empty'] += 1
                    else:
                        sheet_info['formula_with_value'] += 1
                        cs['with_value'] += 1
                        
                        if cs['sample_result'] is None and len(sheet_info['sample_results']) < 5:
                            sheet_info['sample_results'].append({
                                'cell': fc.coordinate,
                                'formula': fc.value[:120],
                                'result': str(result)[:80]
                            })
                            cs['sample_result'] = str(result)[:80]
        
        # Классификация столбцов
        for cl, st in col_stats.items():
            if st['with_value'] == 0:
                sheet_info['dead_columns'].append(cl)
                st['status'] = 'dead'
            elif st['empty'] == 0:
                sheet_info['live_columns'].append(cl)
                st['status'] = 'live'
            else:
                sheet_info['mixed_columns'].append(cl)
                st['status'] = 'mixed'
        
        sheet_info['column_details'] = col_stats
        results[ws_name] = sheet_info
        
        # Вывод
        total = sheet_info['formula_cells']
        active = sheet_info['formula_with_value']
        empty = sheet_info['formula_empty']
        pct = (active/total*100) if total > 0 else 0
        
        print(f"\n{'='*70}")
        print(f"{ws_name}: {total:,} ф-л → {active:,} ({pct:.0f}%) дают результат, "
              f"{empty:,} ({100-pct:.0f}%) пустые")
        print(f"  Столбцы: живые={len(sheet_info['live_columns'])}, "
              f"смешанные={len(sheet_info['mixed_columns'])}, "
              f"мёртвые={len(sheet_info['dead_columns'])}")
        
        if sheet_info['dead_columns']:
            print(f"  ⚠️ Мёртвые: {sheet_info['dead_columns'][:20]}")
        if sheet_info['mixed_columns']:
            print(f"  🔀 Смешанные: {sheet_info['mixed_columns'][:15]}")
        if sheet_info['sample_results']:
            print(f"  Примеры:")
            for s in sheet_info['sample_results'][:3]:
                print(f"    {s['cell']}: {s['result']}")
    
    wb_form.close()
    wb_val.close()
    
    # Итоги
    print(f"\n{'='*70}")
    print("ИТОГО:")
    grand_t = grand_a = grand_e = 0
    all_dead = []
    
    for ws, info in sorted(results.items(), key=lambda x: x[1]['formula_cells'], reverse=True):
        t, a, e = info['formula_cells'], info['formula_with_value'], info['formula_empty']
        grand_t += t; grand_a += a; grand_e += e
        pct = (a/t*100) if t > 0 else 0
        print(f"  {ws:45s} {t:>7,} → {a:>6,} ({pct:3.0f}%)  пустых:{e:>6,}  "
              f"мёртвых стлб:{len(info['dead_columns'])}")
        all_dead.extend([(ws, c) for c in info['dead_columns']])
    
    gp = (grand_a/grand_t*100) if grand_t else 0
    print(f"\n  {'='*70}")
    print(f"  ВСЕГО: {grand_t:,} формул → работают: {grand_a:,} ({gp:.0f}%), "
          f"пустые: {grand_e:,} ({100-gp:.0f}%)")
    print(f"  Мёртвых столбцов: {len(all_dead)}")
    
    with open(os.path.join(OUTPUT_DIR, 'formula_activity.json'), 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    log(f"Результат: {OUTPUT_DIR}/formula_activity.json")

if __name__ == '__main__':
    analyze()
