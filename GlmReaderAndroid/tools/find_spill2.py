"""Ищем НАСТОЯЩИЕ spill-формулы Excel 365: FILTER, UNIQUE, SORT, LET, LAMBDA"""
import openpyxl
import re

PROCESS_FILE = r'c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Обработка.xlsx'

# Настоящие spill-функции (русские + английские)
SPILL_FUNCS = [
    # Английские
    r'\bFILTER\b', r'\bUNIQUE\b', r'\bSORT\b', r'\bSORTBY\b',
    r'\bLET\b', r'\bLAMBDA\b', r'\bMAP\b', r'\bREDUCE\b',
    r'\bBYROW\b', r'\bBYCOL\b', r'\bSCAN\b', r'\bMAKEARRAY\b',
    r'\bTOCOL\b', r'\bTOROW\b', r'\bVSTACK\b', r'\bHSTACK\b',
    r'\bSEQUENCE\b', r'\bTAKE\b', r'\bDROP\b', r'\bCHOOSEROWS\b',
    # Русские
    r'\bФИЛЬТР\b', r'\bУНИК\b', r'\bСОРТ\b', r'\bСЖПРОБЕЛЫ\b',
]

wb = openpyxl.load_workbook(PROCESS_FILE, data_only=False, read_only=True)

found = []
for ws in wb.worksheets:
    ws_name = ws.title
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 502), max_col=min(ws.max_column, 225), values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                for pattern in SPILL_FUNCS:
                    if re.search(pattern, cell.value, re.IGNORECASE):
                        found.append({
                            'sheet': ws_name,
                            'cell': cell.coordinate,
                            'formula': cell.value
                        })
                        break

wb.close()

print(f"Найдено {len(found)} настоящих spill-формул:\n")

by_sheet = {}
for f in found:
    by_sheet.setdefault(f['sheet'], []).append(f)

for ws_name, formulas in sorted(by_sheet.items()):
    # Группируем по шаблону
    by_type = {}
    for f in formulas:
        tpl = re.sub(r'\$?[A-Z]+\$?\d+', '$REF', f['formula'])
        by_type.setdefault(tpl, []).append(f)
    
    print(f"\n--- {ws_name}: {len(formulas)} spill-формул, {len(by_type)} уникальных ---")
    
    for tpl, instances in sorted(by_type.items(), key=lambda x: len(x[1]), reverse=True):
        inst = instances[0]
        formula = inst['formula']
        
        print(f"\n  [{len(instances)} ячеек] {inst['cell']}:")
        
        # Разворачиваем по скобкам
        depth = 0
        out = []
        i = 0
        in_quotes = False
        while i < len(formula):
            ch = formula[i]
            if ch == '"':
                in_quotes = not in_quotes
                out.append(ch)
                i += 1
            elif not in_quotes and ch == '(':
                depth += 1
                out.append('(\n' + '  ' * depth)
                i += 1
            elif not in_quotes and ch == ')':
                depth -= 1
                out.append('\n' + '  ' * depth + ')')
                i += 1
            elif not in_quotes and ch in (';', ','):
                out.append(ch + '\n' + '  ' * depth)
                i += 1
            else:
                j = i
                while j < len(formula) and not (not in_quotes and formula[j] in '();,'):
                    if formula[j] == '"':
                        in_quotes = not in_quotes
                    j += 1
                out.append(formula[i:j])
                i = j
        
        expanded = ''.join(out)
        # Обрезаем если слишком длинная
        lines = expanded.split('\n')
        if len(lines) > 20:
            print('\n'.join(lines[:15]))
            print(f"  ... ({len(lines)-15} строк скрыто)")
        else:
            print(expanded)
