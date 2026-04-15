"""Ищем spill-формулы: FILTER, UNIQUE, SORT, LET, LAMBDA, MAP, ДВССЫЛ+FILTER и т.д."""

import openpyxl
import re

PROCESS_FILE = r"c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Обработка.xlsx"

# Ключевые слова spill-формул
SPILL_KEYWORDS = [
    "FILTER",
    "UNIQUE",
    "SORT",
    "SORTBY",
    "LET",
    "LAMBDA",
    "MAP",
    "REDUCE",
    "SCAN",
    "MAKEARRAY",
    "XLOOKUP",
    "TOCOL",
    "TOROW",
    "WRAPCOLS",
    "WRAPROWS",
    "VSTACK",
    "HSTACK",
    "CHOOSEROWS",
    "TAKE",
    "DROP",
    "ФИЛЬТР",
    "УНИК",
    "СОРТ",
    "СОРТПО",
    "ДВССЫЛ",
    "СЖПРОБЕЛЫ",
    "ДАННЫМИТАБЛ",
    "SEQUENCE",
    "EXPAND",
    "ISOMITTED",
]

wb = openpyxl.load_workbook(PROCESS_FILE, data_only=False, read_only=True)

found = []

for ws in wb.worksheets:
    ws_name = ws.title
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, 502),
        max_col=min(ws.max_column, 225),
        values_only=False,
    ):
        for cell in row:
            if (
                cell.value
                and isinstance(cell.value, str)
                and cell.value.startswith("=")
            ):
                formula_upper = cell.value.upper()
                matches = [kw for kw in SPILL_KEYWORDS if kw.upper() in formula_upper]
                if matches:
                    found.append(
                        {
                            "sheet": ws_name,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "keywords": matches,
                        }
                    )

wb.close()

print(f"Найдено {len(found)} spill-формул:\n")

# Группируем по листам
by_sheet = {}
for f in found:
    by_sheet.setdefault(f["sheet"], []).append(f)

for ws_name, formulas in sorted(by_sheet.items()):
    print(f"\n{'='*80}")
    print(f"{ws_name}: {len(formulas)} spill-формул")
    print(f"{'='*80}")

    # Группируем по типу формулы
    by_type = {}
    for f in formulas:
        tpl = re.sub(r"\$?[A-Z]+\$?\d+", "$REF", f["formula"])
        by_type.setdefault(tpl, []).append(f)

    for tpl, instances in sorted(
        by_type.items(), key=lambda x: len(x[1]), reverse=True
    ):
        inst = instances[0]
        print(f"\n  📐 {len(instances)} ячеек — {', '.join(inst['keywords'])}")

        # Разворачиваем формулу
        formula = inst["formula"]
        if len(formula) > 100:
            # Показываем с отступами
            depth = 0
            out = []
            i = 0
            while i < len(formula):
                if formula[i] == "(":
                    depth += 1
                    out.append("(\n" + "  " * depth)
                    i += 1
                elif formula[i] == ")":
                    depth -= 1
                    out.append("\n" + "  " * depth + ")")
                    i += 1
                elif formula[i] == ";":
                    out.append(";\n" + "  " * depth)
                    i += 1
                elif formula[i] == ",":
                    out.append(",\n" + "  " * depth)
                    i += 1
                else:
                    j = i
                    while j < len(formula) and formula[j] not in "();,":
                        j += 1
                    out.append(formula[i:j])
                    i = j
            print("".join(out))
        else:
            print(f"  {formula}")
