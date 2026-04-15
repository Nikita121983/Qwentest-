"""
Анализ связей между Excel-файлами:
1. Юнтолово 8 корпус 1-22 этажи Квартиры Список измерений edit.xlsx (источник)
2. Обработка.xlsx (обработка данных)
"""

import openpyxl
import json

SOURCE_FILE = r"c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Юнтолово 8 корпус 1-22 этажи Квартиры Список измерений edit.xlsx"
PROCESS_FILE = r"c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Обработка.xlsx"


def read_sheet_headers(ws, max_cols=20):
    """Читает заголовки и первые 3 строки"""
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=False)):
        cells = []
        for c in row[:max_cols]:
            if c.value is not None:
                val_repr = repr(c.value)[:100]
                is_formula = isinstance(c.value, str) and c.value.startswith("=")
                cells.append(
                    {"coord": c.coordinate, "value": val_repr, "formula": is_formula}
                )
        if cells:
            rows.append({"row": i + 1, "cells": cells})
    return rows


def find_formula_references(ws, sheet_names=None):
    """Ищет внешние ссылки на другие листы в формулах"""
    refs = []
    for row in ws.iter_rows(min_row=1, values_only=False):
        for c in row:
            if c.value and isinstance(c.value, str) and c.value.startswith("="):
                # Ищем ссылки на другие листы: 'SheetName'!A1 или SheetName!A1
                import re

                sheet_refs = re.findall(r"'?([^'!]+)'?!", c.value)
                for ref in sheet_refs:
                    refs.append(
                        {
                            "cell": c.coordinate,
                            "formula": c.value[:150],
                            "ref_sheet": ref,
                        }
                    )
    return refs


def analyze_source_file():
    """Анализ файла-источника"""
    print("=" * 80)
    print("ФАЙЛ-ИСТОЧНИК: Юнтолово 8 корпус")
    print("=" * 80)

    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=False, read_only=True)
    result = {}

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        headers = read_sheet_headers(ws, max_cols=30)
        print(f"\n--- {ws_name} ({ws.max_row}x{ws.max_column}) ---")

        # Заголовки (строка 1)
        if headers:
            for cell in headers[0].get("cells", []):
                print(f"  {cell['coord']}: {cell['value']}")

        result[ws_name] = {
            "rows": ws.max_row,
            "cols": ws.max_column,
            "headers": [c["value"] for c in (headers[0]["cells"] if headers else [])],
        }

    wb.close()
    return result


def analyze_process_file():
    """Анализ файла обработки"""
    print("\n" + "=" * 80)
    print("ФАЙЛ ОБРАБОТКИ: Обработка.xlsx")
    print("=" * 80)

    wb = openpyxl.load_workbook(PROCESS_FILE, data_only=False, read_only=True)
    result = {}

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        headers = read_sheet_headers(ws, max_cols=20)
        refs = find_formula_references(ws)

        print(f"\n--- {ws_name} ({ws.max_row}x{ws.max_column}) ---")

        # Заголовки
        if headers:
            for cell in headers[0].get("cells", []):
                print(
                    f"  {cell['coord']}: {cell['value']} {'[FORMULA]' if cell['formula'] else ''}"
                )

        # Внешние ссылки на листы
        if refs:
            # Группируем по ссылкам
            from collections import Counter

            ref_counts = Counter(r["ref_sheet"] for r in refs)
            print("  [Внешние ссылки в формулах]:")
            for ref_sheet, cnt in ref_counts.most_common(10):
                print(f"    -> {ref_sheet}: {cnt} формул")

        has_formulas = sum(
            1
            for row in ws.iter_rows(min_row=1, values_only=False)
            for c in row
            if c.value and isinstance(c.value, str) and c.value.startswith("=")
        )
        print(f"  Формул на листе: {has_formulas}")

        result[ws_name] = {
            "rows": ws.max_row,
            "cols": ws.max_column,
            "headers": [c["value"] for c in (headers[0]["cells"] if headers else [])],
            "formula_count": has_formulas,
            "external_refs": list({r["ref_sheet"] for r in refs}),
        }

    wb.close()
    return result


def find_cross_file_links():
    """Ищем ссылки между файлами через именованные диапазоны и связи"""
    print("\n" + "=" * 80)
    print("ПОИСК СВЯЗЕЙ МЕЖДУ ФАЙЛАМИ")
    print("=" * 80)

    # Ищем в Обработка.xlsx ссылки на внешние файлы
    wb = openpyxl.load_workbook(PROCESS_FILE, data_only=False, read_only=True)

    # Проверяем externalLinks
    if hasattr(wb, "external_links") and wb.external_links:
        print(f"Внешние связи: {wb.external_links}")
    else:
        print("Прямых внешних связей (externalLinks) не найдено")

    # Ищем в формулах упоминания имен из исходного файла
    source_wb = openpyxl.load_workbook(SOURCE_FILE, data_only=False, read_only=True)
    source_sheets = set(source_wb.sheetnames)
    source_wb.close()

    cross_refs = []
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
            for c in row:
                if c.value and isinstance(c.value, str) and c.value.startswith("="):
                    # Ищем ссылки на внешние файлы [файл.xlsx]
                    import re

                    ext_refs = re.findall(r"\[([^\]]+)\]", c.value)
                    for ref in ext_refs:
                        if "Юнтолово" in ref or "edit" in ref:
                            cross_refs.append(
                                {"cell": c.coordinate, "sheet": ws_name, "ref": ref}
                            )

    if cross_refs:
        print(f"Найдено {len(cross_refs)} перекрёстных ссылок:")
        for cr in cross_refs[:10]:
            print(f"  {cr['sheet']}!{cr['cell']} -> {cr['ref']}")
    else:
        print("Прямых ссылок на файл 'Юнтолово...edit.xlsx' не найдено.")
        print("Связь вероятно осуществляется через копирование данных (не формулы).")

    wb.close()


# --- Основной анализ ---
source_info = analyze_source_file()
process_info = analyze_process_file()
find_cross_file_links()

# Сохраняем результат
output = {"source": source_info, "process": process_info}
with open(
    "C:\\Users\\Nik\\SyncthingServiceAcct\\Qwentest\\excel_analysis.json",
    "w",
    encoding="utf-8",
) as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print("\n\nРезультат сохранён в excel_analysis.json")
print("Готово!")
