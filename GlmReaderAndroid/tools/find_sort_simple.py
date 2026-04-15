"""Простой текстовый поиск СОРТ(УНИК"""

import openpyxl

PROCESS_FILE = r"c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Обработка.xlsx"

wb = openpyxl.load_workbook(PROCESS_FILE, data_only=False, read_only=True)

for ws in wb.worksheets:
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, 100),
        max_col=min(ws.max_column, 50),
        values_only=False,
    ):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if (
                    "СОРТ" in cell.value
                    or "УНИК" in cell.value
                    or "ФИЛЬТР" in cell.value
                ):
                    print(f"{ws.title}!{cell.coordinate}:")
                    print(cell.value[:300])
                    print()

wb.close()
