"""Показывает spill-формулы и dynamic ranges на листе Список фильтр"""

import openpyxl

PROCESS_FILE = r"c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Обработка.xlsx"

wb = openpyxl.load_workbook(PROCESS_FILE, data_only=False)
ws = wb["Список фильтр"]

print(f"Список фильтр: {ws.max_row}x{ws.max_column}\n")

for col_num in range(1, min(ws.max_column + 1, 51)):
    col_letter = openpyxl.utils.get_column_letter(col_num)
    formulas = []

    for row in range(1, 10):
        cell = ws.cell(row=row, column=col_num)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
            formulas.append((row, cell.value))

    if formulas:
        print(f"--- Столбец {col_letter} ({len(formulas)} формул в первых строках) ---")
        for row, formula in formulas:
            short = formula[:120]
            is_dynamic = any(
                kw in formula
                for kw in [
                    "INDIRECT",
                    "OFFSET",
                    "COUNTA",
                    "CONCAT",
                    "FILTER",
                    "#",
                    ":$",
                ]
            )
            marker = " 📐" if is_dynamic else ""
            print(f"  {col_letter}{row}: {short}{marker}")
        print()

wb.close()
