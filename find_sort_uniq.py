"""Ищем формулы с СОРТ(УНИК(..."""
import openpyxl
import re

PROCESS_FILE = r'c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Обработка.xlsx'

wb = openpyxl.load_workbook(PROCESS_FILE, data_only=False, read_only=True)

patterns_to_find = [
    r'СОРТ\s*\(\s*УНИК',
    r'УНИК\s*\(',
    r'СОРТ\s*\(',
    r'СЖПРОБЕЛЫ\s*\(',
    r'ФИЛЬТР\s*\(',
    r'ДВССЫЛ\s*\(',
]

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 502), max_col=min(ws.max_column, 225), values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                for pat in patterns_to_find:
                    if re.search(pat, cell.value):
                        print(f"{ws.title}!{cell.coordinate}:")
                        # Разворачиваем
                        depth = 0
                        out = []
                        i = 0
                        in_q = False
                        while i < len(cell.value) and i < 500:
                            ch = cell.value[i]
                            if ch == '"':
                                in_q = not in_q
                                out.append(ch)
                            elif not in_q and ch == '(':
                                depth += 1
                                out.append('(\n' + '  ' * depth)
                            elif not in_q and ch == ')':
                                depth -= 1
                                out.append('\n' + '  ' * depth + ')')
                            elif not in_q and ch in (';', ','):
                                out.append(ch + '\n' + '  ' * depth)
                            else:
                                out.append(ch)
                            i += 1
                        print(''.join(out))
                        print()
                        break

wb.close()
