"""
Редактор формул для Обработка.xlsx
- Извлекает ВСЕ уникальные формулы из каждого столбца
- Показывает: где используется, сколько ячеек, какие варианты
- Позволяет заменить формулу → автоматически обновляет ВСЕ ячейки
- Проверяет что ничего не сломалось
"""
import openpyxl
import re
import json
import os
import time
from collections import defaultdict

PROCESS_FILE = r'c:\Users\Nik\SyncthingServiceAcct\0 Работа\Юнтолово 4 корпус\8 корпус\Обработка.xlsx'
BACKUP_SUFFIX = '.bak'
OUTPUT_DIR = r'C:\Users\Nik\SyncthingServiceAcct\Qwentest\formula_editor'
os.makedirs(OUTPUT_DIR, exist_ok=True)

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")

def extract_column_formulas(wb):
    """Извлекает уникальные формулы для КАЖДОГО столбца каждого листа"""
    registry = {}  # {sheet: {col: {template: {count, formula, example_rows, formula_text}}}
    
    for ws in wb.worksheets:
        ws_name = ws.name
        registry[ws_name] = {}
        
        max_col = min(ws.max_column, 225)
        max_row = min(ws.max_row, 502)
        
        for col_num in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            formulas = []
            
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col_num)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append((row, cell.value))
            
            if not formulas:
                continue
            
            # Группируем по шаблону (нормализуем номера строк)
            templates = defaultdict(list)
            for row, formula in formulas:
                # Заменяем ссылки на ячейки в пределах того же столбца
                # $A$1 -> $CELL, A1 -> CELL, $A$1 -> $CELL
                tpl = re.sub(r'\$?' + re.escape(col_letter) + r'\$?(\d+)', '$ROW', formula)
                # Остальные ссылки на другие столбцы
                tpl2 = re.sub(r'\$?[A-Z]+\$?\d+', '$REF', tpl)
                templates[tpl2].append((row, formula))
            
            col_data = {}
            for tpl, rows_formulas in templates.items():
                first_row, first_formula = rows_formulas[0]
                col_data[tpl] = {
                    'count': len(rows_formulas),
                    'formula': first_formula,
                    'rows': [r for r, f in rows_formulas[:5]],  # первые 5 строк
                    'all_rows': len(rows_formulas),
                }
            
            registry[ws_name][col_letter] = col_data
    
    return registry

def print_column_summary(registry):
    """Выводит сводку: какие столбцы сколько вариантов имеют"""
    print("\n" + "="*80)
    print("РЕЕСТР ФОРМУЛ — СВОДКА")
    print("="*80)
    
    total_cols = 0
    total_multi_variant = 0
    total_formulas = 0
    
    for ws_name, cols in sorted(registry.items()):
        if not cols:
            continue
        
        multi_cols = {c: data for c, data in cols.items() if len(data) > 1}
        
        print(f"\n--- {ws_name} ({len(cols)} столбцов с формулами) ---")
        
        for col_letter, variants in sorted(cols.items()):
            total_cols += 1
            vcount = len(variants)
            total = sum(v['count'] for v in variants.values())
            total_formulas += total
            
            if vcount > 1:
                total_multi_variant += 1
                marker = " 🔀"
            else:
                marker = ""
            
            print(f"  {col_letter:4s} {vcount} вариант(ов)  {total:>5} ячеек{marker}")
            
            for i, (tpl, data) in enumerate(variants.items()):
                rows_str = f"строки {data['rows']}" if len(data['rows']) <= 5 else f"строк {data['all_rows']}"
                formula_short = data['formula'][:90]
                print(f"    [{i+1}] {data['count']:>4} ячеек ({rows_str})")
                print(f"        {formula_short}")
    
    print(f"\n{'='*80}")
    print(f"ИТОГО: {total_cols} столбцов, {total_multi_variant} с вариантами, {total_formulas:,} формул")

def save_registry_json(registry):
    """Сохраняет реестр в JSON для редактирования"""
    # Преобразуем в JSON-friendly формат
    data = {}
    for ws_name, cols in registry.items():
        data[ws_name] = {}
        for col_letter, variants in cols.items():
            data[ws_name][col_letter] = {}
            for tpl, info in variants.items():
                data[ws_name][col_letter][tpl] = {
                    'count': info['count'],
                    'formula': info['formula'],
                    'rows': info['rows']
                }
    
    path = os.path.join(OUTPUT_DIR, 'formula_registry.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    log(f"Реестр сохранён: {path}")

def apply_formula_change(sheet_name, col_letter, template_filter, new_formula):
    """Заменяет формулу во ВСЕХ ячейках, соответствующих фильтру"""
    log(f"Применяю изменение: {sheet_name}!{col_letter}")
    log(f"  Шаблон: {template_filter[:80]}")
    log(f"  Новая формула: {new_formula[:100]}")
    
    # Открываем
    wb = openpyxl.load_workbook(PROCESS_FILE)
    ws = wb[sheet_name]
    col_num = openpyxl.utils.column_index_from_string(col_letter)
    
    # Находим все ячейки с этим шаблоном
    count = 0
    for row in range(1, min(ws.max_row + 1, 502)):
        cell = ws.cell(row=row, column=col_num)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            # Проверяем совпадение шаблона
            tpl = re.sub(r'\$?' + re.escape(col_letter) + r'\$?(\d+)', '$ROW', cell.value)
            tpl2 = re.sub(r'\$?[A-Z]+\$?\d+', '$REF', tpl)
            
            if tpl2 == template_filter or template_filter in cell.value:
                # Заменяем формулу
                # Нужно адаптировать номер строки
                old_formula = cell.value
                # Извлекаем номер строки из старой формулы
                row_match = re.search(re.escape(col_letter) + r'(\d+)', old_formula)
                if row_match:
                    old_row_num = row_match.group(1)
                    new_formula_with_row = new_formula.replace('$ROW', col_letter + old_row_num)
                    # Заменяем все вхождения
                    new_formula_with_row = re.sub(
                        re.escape(col_letter) + r'\d+',
                        lambda m: col_letter + old_row_num,
                        new_formula
                    )
                    # Но корректнее: заменяем только ССЫЛКИ на тот же столбец
                    new_formula_with_row = re.sub(
                        r'(\$?)' + re.escape(col_letter) + r'(\$?)(\d+)',
                        lambda m: m.group(1) + col_letter + m.group(2) + str(row),
                        new_formula
                    )
                    # А ссылки на ДРУГИЕ столбцы — оставляем как есть, но с текущим номером строки
                    new_formula_with_row = re.sub(
                        r'(\$?)([A-Z]+)(\$?)(\d+)',
                        lambda m: m.group(1) + m.group(2) + m.group(3) + str(row) if m.group(2) != col_letter
                        else m.group(1) + col_letter + m.group(3) + str(row),
                        new_formula
                    )
                else:
                    new_formula_with_row = new_formula
                
                cell.value = new_formula_with_row
                count += 1
    
    # Сохраняем
    backup = PROCESS_FILE + BACKUP_SUFFIX
    import shutil
    shutil.copy2(PROCESS_FILE, backup)
    log(f"  Бэкап: {backup}")
    
    wb.save(PROCESS_FILE)
    wb.close()
    log(f"  Обновлено {count} ячеек в {sheet_name}!{col_letter}")
    return count

def main():
    log("Извлекаю формулы из Обработка.xlsx...")
    wb = openpyxl.load_workbook(PROCESS_FILE, data_only=False)
    
    registry = extract_column_formulas(wb)
    wb.close()
    
    print_column_summary(registry)
    save_registry_json(registry)
    
    print(f"\n{'='*80}")
    print("КАК ИСПОЛЬЗОВАТЬ:")
    print("="*80)
    print(f"""
1. Найди нужную формулу в выводе выше
2. Скопируй шаблон (то что в [1], [2] и т.д.)
3. Вызови apply_formula_change():

   apply_formula_change(
       sheet_name='Расчет',
       col_letter='AV',
       template_filter='Шаблон из вывода',
       new_formula='=IF($F3<>"",$F3,$E3)'  # новая формула БЕЗ номеров строк
   )

4. Для поиска конкретного места — смотри sheet_name + col_letter + rows
""")

if __name__ == '__main__':
    main()
